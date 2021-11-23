"""
模块介绍：
模块名称：dictionary.py      模块地址：~project~/rout_lib/dictionary.py
模块用途：与主程序蓝图链接，显示部分路由配置，此处路由皆用于与substitute.vocabulary/words/poetry数据库连接交互并根据前端关键词搜索数据并返回
模块运行环境配置：app（主函数）自定义库 flask库 openpyxl库
模块版本：3.12               最近修改：2021/5/21
"""

# 部署运行环境 引用外部库
import app                                      # 自定义库：主函数
from flask import Blueprint, request, jsonify   # flask库：flask框架本身
from openpyxl import load_workbook              # openpyxl库：与xlsx表格文件交互

# flask蓝图函数蓝图链接主函数与模块路由
dictionary = Blueprint('dictionary', __name__)

# 具体路由配置


@dictionary.route('/dictionary/dictionary_vocabulary_search', methods=['POST'])
def dictionary_vocabulary_search():
    """
    函数名称:   词汇关键字查询
    函数作用:   使用关键字查询词汇
    传递参数:   str( kwd ) --> 搜索用关键字
    返回值：    json.array --> 词汇名称组成的json数组
    """
    if request.method == 'POST':
        kwd = str(request.form['kwd'])
        app.cursor.execute("SELECT id FROM keyword WHERE type REGEXP '{kwd}';".format(kwd=kwd))
        app.db.commit()
        ex_kwd_id = app.cursor.fetchall()
        kwd_id = int(ex_kwd_id[0][0])
        app.cursor.execute("SELECT vname FROM vocabulary WHERE vkeywords REGEXP '{kwd_id}';".format(kwd_id=kwd_id))
        app.db.commit()
        data = app.cursor.fetchall()
        return jsonify(data)


@dictionary.route('/dictionary/dictionary_poetry_search', methods=['POST'])
def dictionary_poetry_search():
    """
    函数名称：   诗歌关键字查询
    函数作用：   使用关键字查询诗歌
    传递参数：   str( kwd ) --> 搜索用关键字
    返回值：    json.array --> 诗歌内容组成的json数组
    """
    if request.method == 'POST':
        kwd = str(request.form['kwd'])
        app.cursor.execute("SELECT id FROM keyword WHERE type REGEXP '{kwd}';".format(kwd=kwd))
        app.db.commit()
        ex_kwd_id = app.cursor.fetchall()
        kwd_id = int(ex_kwd_id[0][0])
        app.cursor.execute("SELECT id FROM poetry WHERE pkeywords REGEXP '{kwd_id}';".format(kwd_id=kwd_id))
        app.db.commit()
        id_data = app.cursor.fetchall()
        wb = load_workbook(filename=r'F:\成品\程序\辞藻替换\后端软件\sheets\poetry.xlsx')
        sheet = wb['Sheet1']
        dps_data = []
        for num in id_data:
            num = int(num[0])
            dps_data.append(sheet.cell(row=num, column=2).value)
        return jsonify(dps_data)


@dictionary.route('/dictionary/dictionary_words_search', methods=['POST'])
def dictionary_words_search():
    """
    函数名称：   名人名言关键字查询
    函数作用：   使用关键字查询名人名言
    传递参数：   str( kwd ) --> 搜索用关键字
    返回值：    json.array --> 名人名言内容组成的json数组
    """
    if request.method == 'POST':
        kwd = str(request.form['kwd'])
        app.cursor.execute("SELECT id FROM keyword WHERE type REGEXP '{kwd}';".format(kwd=kwd))
        app.db.commit()
        ex_kwd_id = app.cursor.fetchall()
        kwd_id = int(ex_kwd_id[0][0])
        app.cursor.execute("SELECT id FROM words WHERE wkeywords REGEXP '{kwd_id}';".format(kwd_id=kwd_id))
        app.db.commit()
        id_data = app.cursor.fetchall()
        wb = load_workbook(filename=r'F:\成品\程序\辞藻替换\后端软件\sheets\words.xlsx')
        sheet = wb['Sheet1']
        dps_data = []
        for num in id_data:
            num = int(num[0])
            dps_data.append(sheet.cell(row=num, column=1).value)
        return jsonify(dps_data)


@dictionary.route('/dictionary/dictionary_test', methods=['GET'])
def dictionary_test():
    """
    测试函数
    """
    if request.method == 'GET':
        return 'Hello World!\n The function of dictionary has normal operate ~ ~ ~'
