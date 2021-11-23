"""
模块介绍：
模块名称：login.py      模块地址：~project~/rout_lib/login.py
模块用途：与主程序蓝图链接，显示部分路由配置，此处路由皆用于与substitute.login数据库连接交互并进行用户注册登录等操作
模块运行环境配置：app（主函数）自定义库 flask库
模块版本：1.3               最近修改：2021/5/21
"""

# 部署运行环境 引用外部库
import app                                      # 自定义库：主函数
from flask import Blueprint, request, jsonify   # flask库：flask框架本身
from openpyxl import load_workbook              # openpyxl库：与xlsx表格文件交互

# flask蓝图函数蓝图链接主函数与模块路由
login = Blueprint('login', __name__)

# 具体路由配置


@login.route('/login/login_register', methods=['POST'])
def login_register():
    """
    函数名称:   注册表注册用户
    函数作用:   向用户表中注册新用户
    传递参数:   str( username ) --> 新用户用户名
              str( password ) --> 新用户密码
    返回值:    注册成功时返回：    json.array --> 用户ID
              失败时返回 -1
    """
    if request.method == 'POST':
        username = str(request.form['username'])
        password = str(request.form['password'])
        try:
            app.cursor.execute("INSERT INTO login(username,`password`) VALUES ('%s','%s');" % (username, password))
            app.db.commit()
            app.cursor.execute("SELECT id FROM login WHERE username REGEXP '{username}';".format(username=username))
            app.db.commit()
            data = app.cursor.fetchall()
            return jsonify(data)
        except:
            app.db.rollback()
            return "-1"


@login.route('/login/login_logon', methods=['POST'])  # 登录
def login_logon():
    """
    函数名称:   用户登录
    函数作用:   在注册表中检查用户名及密码
    传递参数:   str( username ) --> 用户用户名
               str( password ) --> 用户密码
    返回值:    登录成功时返回：1
                用户名或密码错误时返回 -1
                用户名不存在时返回 0
    """
    if request.method == 'POST':
        username = str(request.form['username'])
        password = str(request.form['password'])
        app.cursor.execute("SELECT password FROM login WHERE username REGEXP '{username}';".format(username=username))
        app.db.commit()
        ex_true_password = app.cursor.fetchall()
        if ex_true_password == []:
            true_password = str(ex_true_password[0][0])
            if true_password == password:
                return "1"
            else:
                return "-1"
        else:
            return "0"


@login.route('/login/login_username_change', methods=['POST'])  # 修改用户名
def login_username_change():
    if request.method == 'POST':
        user_id = str(request.form['user_id'])
        new_username = str(request.form['new_username'])
        try:
            app.cursor.execute("UPDATE login SET username = '%s' WHERE id='%s';" %(new_username,user_id))
            app.db.commit()
            return "1"
        except:
            return "0"


@login.route('/login/login_passwoord_change', methods=['POST']) # 修改密码
def login_passwoord_change():
    if request.method == 'POST':
        username = str(request.form['username'])
        new_password = str(request.form['new_password'])
        try:
            app.cursor.execute("UPDATE login SET password = '%s' WHERE username='%s';" %(new_password,username))
            app.db.commit()
            return "1"
        except:
            return "0"


@login.route('/login/login_favorate_add', methods=['POST']) # 添加收藏
def login_favorate_add():
    if request.method == 'POST':
        user_id = str(request.form['user_id'])
        data_id = str(request.form['data_id'])
        dic_type = int(request.form['dic_type'])
        flag = 0
        if dic_type == 1:
            app.cursor.execute("SELECT favorate_v FROM login WHERE id REGEXP '%s';" % user_id)
            app.db.commit()
            ex_favorite_id = app.cursor.fetchall()
            favorite_id = str(ex_favorite_id[0][0])
            favorite_list = favorite_id.split(';')
            for num in favorite_list:
                if data_id == num:
                    flag = 1
                    break
            if flag == 1:
                return "0"
            else:
                app.cursor.execute("UPDATE login SET favorate_v = CONCAT(favorate_v,'%s;') WHERE id='%s';" %(data_id,user_id))
                app.db.commit()
                return "1"
        elif dic_type == 2:
            app.cursor.execute("SELECT favorate_w FROM login WHERE id REGEXP '%s';" % user_id)
            app.db.commit()
            ex_favorite_id = app.cursor.fetchall()
            favorite_id = str(ex_favorite_id[0][0])
            favorite_list = favorite_id.split(';')
            for num in favorite_list:
                if data_id == num:
                    flag = 1
                    break
            if flag == 1:
                return "0"
            else:
                app.cursor.execute(
                    "UPDATE login SET favorate_w = CONCAT(favorate_v,'%s;') WHERE id='%s';" % (data_id, user_id))
                app.db.commit()
                return "1"
        elif dic_type == 3:
            app.cursor.execute("SELECT favorate_p FROM login WHERE id REGEXP '%s';" % user_id)
            app.db.commit()
            ex_favorite_id = app.cursor.fetchall()
            favorite_id = str(ex_favorite_id[0][0])
            favorite_list = favorite_id.split(';')
            for num in favorite_list:
                if data_id == num:
                    flag = 1
                    break
            if flag == 1:
                return "0"
            else:
                app.cursor.execute(
                    "UPDATE login SET favorate_p = CONCAT(favorate_v,'%s;') WHERE id='%s';" % (data_id, user_id))
                app.db.commit()
                return "1"


@login.route('/login/login_favorate_show', methods=['POST']) # 查看收藏
def login_favorate_show():
    if request.method == 'POST':
        user_id = str(request.form['user_id'])
        dic_type = int(request.form['dic_type'])
        try:
            if dic_type == 1:
                app.cursor.execute("SELECT favorate_v FROM login WHERE id REGEXP '%s';" % user_id)
                app.db.commit()
                ex_data_id = app.cursor.fetchall()
                data_id = str(ex_data_id[0][0])
                data_list = data_id.split(';')
                data = []
                for num in data_list:
                    if num != None:
                        app.cursor.execute("SELECT vname FROM vocabulary WHERE id='%s';" % num)
                        app.db.commit()
                        ex_data = app.cursor.fetchall()
                        data.extend(ex_data)
                return jsonify(data)
            elif dic_type == 2:
                app.cursor.execute("SELECT favorate_w FROM login WHERE id REGEXP '%s';" % user_id)
                app.db.commit()
                ex_data_id = app.cursor.fetchall()
                data_id = str(ex_data_id[0][0])
                data_list = data_id.split(';')
                wb = load_workbook(filename=r'F:\成品\程序\辞藻替换\后端软件\sheets\words.xlsx')
                sheet = wb['Sheet1']
                data = []
                for num in data_list:
                    num = int(num[0])
                    data.append(sheet.cell(row=num, column=2).value)
                return jsonify(data)
            elif dic_type == 3:
                app.cursor.execute("SELECT favorate_p FROM login WHERE id REGEXP '%s';" % user_id)
                app.db.commit()
                ex_data_id = app.cursor.fetchall()
                data_id = str(ex_data_id[0][0])
                data_list = data_id.split(';')
                wb = load_workbook(filename=r'F:\成品\程序\辞藻替换\后端软件\sheets\poetry.xlsx')
                sheet = wb['Sheet1']
                data = []
                for num in data_list:
                    num = int(num[0])
                    data.append(sheet.cell(row=num, column=2).value)
                return jsonify(data)
        except:
            return "-1"


@login.route('/login/login_test', methods=['GET'])
def login_test():
    """
    测试函数
    """
    if request.method == 'GET':
        return 'Hello World!\n The function of login has normal operate ~ ~ ~'
